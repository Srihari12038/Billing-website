from openpyxl import Workbook, load_workbook

from .models import Category, Product


def export_products_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["Name", "SKU", "Barcode", "Category", "Cost Price", "Selling Price", "GST", "HSN", "Current Stock", "Minimum Stock"])
    for product in Product.objects.select_related("category"):
        ws.append([product.name, product.sku, product.barcode, product.category.name, product.cost_price, product.selling_price, product.gst_rate, product.hsn_code, product.current_stock, product.minimum_stock])
    return wb


def import_products_workbook(file_obj):
    wb = load_workbook(file_obj)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        category, _ = Category.objects.get_or_create(name=row[3] or "General")
        Product.objects.update_or_create(
            sku=row[1],
            defaults={
                "name": row[0],
                "barcode": row[2] or "",
                "category": category,
                "cost_price": row[4] or 0,
                "selling_price": row[5] or 0,
                "gst_rate": row[6] or 0,
                "hsn_code": row[7] or "",
                "current_stock": row[8] or 0,
                "minimum_stock": row[9] or 0,
            },
        )
        count += 1
    return count
